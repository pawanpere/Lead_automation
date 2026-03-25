#!/usr/bin/env python3
"""
Lead Qualification & Outreach Automation
=========================================
Agentic pipeline that:
1. Reads ecommerce brands from an Excel sheet
2. Fires ScraperAgent (Playwright) to scrape website + take product screenshot
3. Fires QualifyAgent (Claude CLI) to qualify the lead
4. Fires EmailAgent (Claude CLI + PlusVibe API) to draft and send personalized email

Usage:
    python3 main.py                          # Process data/leads.xlsx
    python3 main.py --file path/to/file.xlsx # Custom Excel file
    python3 main.py --concurrency 5          # Max 5 brands at once
    python3 main.py --dry-run                # Qualify only, don't send emails
"""

import asyncio
import argparse
import json
import logging
import os
import sys
from datetime import datetime

from openpyxl import load_workbook

from agents import ScraperAgent, QualifyAgent, EmailAgent

# --- Logging setup ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(
            os.path.join(os.path.dirname(__file__), "output", "logs", "run.log"),
            mode="a",
        ),
    ],
)
logger = logging.getLogger("orchestrator")

BASE_DIR = os.path.dirname(__file__)
OUTPUT_FILE = os.path.join(BASE_DIR, "output", "results.json")


def read_excel(file_path: str):
    """Read leads from Excel file. Expected columns: Brand Name, URL, Email, Product Page Link."""
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # First row is header — normalize column names
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    # Map common column name variations
    col_map = {}
    for i, h in enumerate(headers):
        if "brand" in h or "name" in h:
            col_map["brand_name"] = i
        elif "url" in h and "product" not in h:
            col_map["url"] = i
        elif "email" in h or "mail" in h:
            col_map["email"] = i
        elif "product" in h:
            col_map["product_url"] = i

    leads = []
    for row in rows[1:]:
        if not row or not any(row):
            continue

        lead = {
            "brand_name": str(row[col_map.get("brand_name", 0)] or "").strip(),
            "url": str(row[col_map.get("url", 1)] or "").strip(),
            "email": str(row[col_map.get("email", 2)] or "").strip(),
            "product_url": str(row[col_map.get("product_url", 3)] or "").strip(),
        }

        # Skip rows with missing essential data
        if lead["brand_name"] and lead["url"]:
            # Ensure URLs have protocol
            if lead["url"] and not lead["url"].startswith("http"):
                lead["url"] = "https://" + lead["url"]
            if lead["product_url"] and not lead["product_url"].startswith("http"):
                lead["product_url"] = "https://" + lead["product_url"]
            leads.append(lead)

    wb.close()
    logger.info(f"Loaded {len(leads)} leads from {file_path}")
    return leads


async def process_brand(lead: dict, dry_run: bool = False) -> dict:
    """Full agent pipeline for a single brand."""
    brand = lead["brand_name"]
    result = {
        "brand_name": brand,
        "url": lead["url"],
        "email": lead["email"],
        "timestamp": datetime.now().isoformat(),
        "scrape": None,
        "qualification": None,
        "email_result": None,
        "status": "pending",
    }

    # --- Agent 1: Scrape ---
    logger.info(f"{'='*50}")
    logger.info(f"BRAND: {brand}")
    logger.info(f"{'='*50}")
    logger.info(f"[1/3] Firing ScraperAgent for {brand}...")

    scrape_result = await ScraperAgent.run(
        brand_name=brand,
        url=lead["url"],
        product_url=lead["product_url"],
    )
    result["scrape"] = {
        "success": scrape_result["success"],
        "screenshot": scrape_result["screenshot_path"],
        "error": scrape_result.get("error"),
    }

    if not scrape_result["success"]:
        result["status"] = "scrape_failed"
        logger.error(f"[ScraperAgent] FAILED for {brand}: {scrape_result.get('error')}")
        return result

    # --- Agent 2: Qualify ---
    logger.info(f"[2/3] Firing QualifyAgent for {brand}...")

    qualification = await QualifyAgent.run(
        brand_name=brand,
        url=lead["url"],
        scraped_data=scrape_result["scraped_data"],
    )
    result["qualification"] = qualification

    if not qualification.get("qualified"):
        result["status"] = "not_qualified"
        logger.info(
            f"[QualifyAgent] {brand} NOT QUALIFIED "
            f"(score: {qualification.get('score', '?')}, "
            f"reasons: {qualification.get('reasons', [])})"
        )
        return result

    logger.info(
        f"[QualifyAgent] {brand} QUALIFIED! "
        f"(score: {qualification.get('score')}, niche: {qualification.get('niche')})"
    )

    # --- Agent 3: Email ---
    if dry_run:
        result["status"] = "qualified_dry_run"
        logger.info(f"[3/3] DRY RUN - Skipping email for {brand}")
        return result

    if not lead["email"]:
        result["status"] = "qualified_no_email"
        logger.warning(f"[3/3] No email address for {brand}, skipping send")
        return result

    logger.info(f"[3/3] Firing EmailAgent for {brand}...")

    email_result = await EmailAgent.run(
        brand_name=brand,
        email=lead["email"],
        url=lead["url"],
        niche=qualification.get("niche", "ecommerce"),
        tiktok_angle=qualification.get("tiktok_angle", ""),
        product_count=qualification.get("product_count_estimate", "unknown"),
        screenshot_path=scrape_result["screenshot_path"],
        site_description=scrape_result["scraped_data"].get("page_text", "")[:500],
    )
    result["email_result"] = email_result
    result["status"] = "email_sent" if email_result.get("sent") else "email_failed"

    return result


async def main():
    parser = argparse.ArgumentParser(description="Lead Qualification & Outreach Automation")
    parser.add_argument(
        "--file",
        default=os.path.join(BASE_DIR, "data", "leads.xlsx"),
        help="Path to Excel file with leads",
    )
    parser.add_argument(
        "--concurrency",
        type=int,
        default=3,
        help="Max number of brands to process concurrently (default: 3)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Qualify leads but don't send emails",
    )
    args = parser.parse_args()

    # Ensure output directories exist
    os.makedirs(os.path.join(BASE_DIR, "output", "logs"), exist_ok=True)
    os.makedirs(os.path.join(BASE_DIR, "screenshots"), exist_ok=True)

    logger.info("=" * 60)
    logger.info("LEAD AUTOMATION PIPELINE STARTED")
    logger.info(f"Excel file: {args.file}")
    logger.info(f"Concurrency: {args.concurrency}")
    logger.info(f"Dry run: {args.dry_run}")
    logger.info("=" * 60)

    # Read leads
    leads = read_excel(args.file)
    if not leads:
        logger.error("No leads found in Excel file. Exiting.")
        return

    # Process all brands concurrently with semaphore
    semaphore = asyncio.Semaphore(args.concurrency)
    all_results = []

    async def limited_process(lead):
        async with semaphore:
            return await process_brand(lead, dry_run=args.dry_run)

    results = await asyncio.gather(
        *[limited_process(lead) for lead in leads],
        return_exceptions=True,
    )

    for r in results:
        if isinstance(r, Exception):
            logger.error(f"Brand pipeline exception: {r}")
            all_results.append({"status": "exception", "error": str(r)})
        else:
            all_results.append(r)

    # Save results
    with open(OUTPUT_FILE, "w") as f:
        json.dump(all_results, f, indent=2, default=str)

    # Print summary
    logger.info("")
    logger.info("=" * 60)
    logger.info("PIPELINE COMPLETE — SUMMARY")
    logger.info("=" * 60)

    total = len(all_results)
    qualified = sum(1 for r in all_results if r.get("status") in ("email_sent", "email_failed", "qualified_dry_run", "qualified_no_email"))
    sent = sum(1 for r in all_results if r.get("status") == "email_sent")
    failed = sum(1 for r in all_results if "failed" in r.get("status", ""))

    logger.info(f"Total brands processed:  {total}")
    logger.info(f"Qualified leads:         {qualified}")
    logger.info(f"Emails sent:             {sent}")
    logger.info(f"Failed:                  {failed}")
    logger.info(f"Results saved to:        {OUTPUT_FILE}")
    logger.info("=" * 60)


if __name__ == "__main__":
    asyncio.run(main())
