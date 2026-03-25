#!/usr/bin/env python3
"""Split merged leads into two sheets: with product URL and without."""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

INPUT = "/Users/prakashtupe/lead-automation/output/merged_leads.xlsx"
OUTPUT = "/Users/prakashtupe/lead-automation/output/merged_leads_filtered.xlsx"

print(f"Reading: {INPUT}")
wb_src = load_workbook(INPUT, read_only=True)
ws_src = wb_src.active

rows = list(ws_src.iter_rows(values_only=True))
headers = rows[0]
data = rows[1:]

# Find column indices
header_list = list(headers)
product_url_col = header_list.index("Latest Product URL")
product_title_col = header_list.index("Latest Product Title")

print(f"Total rows: {len(data)}")
print(f"Product URL column: {product_url_col}")
print(f"Product Title column: {product_title_col}")

# Split
with_product = []
without_product = []

for row in data:
    product_url = str(row[product_url_col] or "").strip()
    product_title = str(row[product_title_col] or "").strip()

    if product_url and product_url != "None":
        with_product.append(row)
    else:
        without_product.append(row)

print(f"\nWith product URL:    {len(with_product)}")
print(f"Without product URL: {len(without_product)}")

# Create new workbook with two sheets
wb = Workbook()

# --- Sheet 1: with_product_url ---
ws1 = wb.active
ws1.title = "with_product_url"

header_fill = PatternFill(start_color="16a34a", end_color="16a34a", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=10)

ws1.append(list(headers))
for col in range(1, len(headers) + 1):
    cell = ws1.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

for row in with_product:
    ws1.append(list(row))

ws1.freeze_panes = "A2"

# --- Sheet 2: without_product_url ---
ws2 = wb.create_sheet("without_product_url")

header_fill2 = PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid")

ws2.append(list(headers))
for col in range(1, len(headers) + 1):
    cell = ws2.cell(row=1, column=col)
    cell.fill = header_fill2
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

for row in without_product:
    ws2.append(list(row))

ws2.freeze_panes = "A2"

# Auto-width both sheets
for ws in [ws1, ws2]:
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value or "")
            max_len = max(max_len, min(len(val), 40))
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

wb_src.close()
wb.save(OUTPUT)

print(f"\nSaved: {OUTPUT}")
print(f"  Sheet 'with_product_url':    {len(with_product)} rows (green header)")
print(f"  Sheet 'without_product_url': {len(without_product)} rows (red header)")
