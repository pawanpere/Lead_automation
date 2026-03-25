#!/usr/bin/env python3
"""Generate CSV with screenshot file paths as attachments for PlusVibe."""
import os
import csv
from openpyxl import load_workbook

OUTPUT_DIR = "/Users/prakashtupe/lead-automation/output"
SCREENSHOTS_DIR = "/Users/prakashtupe/lead-automation/screenshots"

# Find latest Excel
excel_files = sorted(
    [f for f in os.listdir(OUTPUT_DIR) if f.startswith("outreach_") and f.endswith(".xlsx")],
    reverse=True,
)
excel_path = os.path.join(OUTPUT_DIR, excel_files[0])
print(f"Reading: {excel_path}")

wb = load_workbook(excel_path, read_only=True)

# Read "Ready to Send" sheet
ws = wb["Ready to Send"]
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]
data = rows[1:]

# Read "All Results" for extra data
ws2 = wb["All Results"]
all_rows = list(ws2.iter_rows(values_only=True))
all_headers = all_rows[0]
all_data = all_rows[1:]

# Build lookup by brand name
all_lookup = {}
for row in all_data:
    brand = row[0]
    all_lookup[brand] = {
        "url": row[1],
        "score": row[4],
        "niche": row[6],
        "products": row[8],
        "est_sales": row[9],
        "tiktok": row[10],
        "country": row[11],
    }

# Generate CSV
csv_path = os.path.join(OUTPUT_DIR, "outreach_ready.csv")

with open(csv_path, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)

    writer.writerow([
        "brand_name",
        "to_email",
        "contact_name",
        "subject",
        "html_body",
        "screenshot_url",
        "screenshot_file",
        "website_url",
        "niche",
        "score",
        "products",
        "tiktok_followers",
        "country",
        "status",
    ])

    count = 0
    for row in data:
        brand_name = row[0] or ""
        to_email = row[1] or ""
        contact_name = row[2] or ""
        subject = row[3] or ""
        html_body = row[4] or ""
        screenshot_url = row[5] or ""
        status = row[8] if len(row) > 8 else ""

        if not to_email or not subject:
            continue

        # Find screenshot file
        safe_name = brand_name.lower().replace(" ", "_").replace("/", "_")
        screenshot_file = os.path.join(SCREENSHOTS_DIR, f"{safe_name}.png")
        if not os.path.exists(screenshot_file):
            screenshot_file = ""

        # Get extra data
        extra = all_lookup.get(brand_name, {})

        writer.writerow([
            brand_name,
            to_email,
            contact_name,
            subject,
            html_body,
            screenshot_url,
            screenshot_file,
            extra.get("url", ""),
            extra.get("niche", ""),
            extra.get("score", ""),
            extra.get("products", ""),
            extra.get("tiktok", ""),
            extra.get("country", ""),
            status,
        ])
        count += 1

wb.close()

print(f"\nCSV saved: {csv_path}")
print(f"Emails ready: {count}")
print(f"\nColumns:")
print("  - screenshot_url  → Cloudinary hosted URL (embedded in html_body as <img>)")
print("  - screenshot_file → Local file path for attachment")
print("  - html_body       → Full email with screenshot inline")
