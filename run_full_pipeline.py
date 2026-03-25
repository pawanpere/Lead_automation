#!/usr/bin/env python3
"""
Full Pipeline: Store Leads → Screenshot → Upload → Qualify → Draft Email → Excel
=================================================================================
Produces an Excel file ready for PlusVibe sending with:
- Brand info, email, contact name
- Hosted screenshot URL (Cloudinary)
- Personalized email subject + body (with screenshot embedded inline)
"""
import asyncio
import sys
import os
import json
import logging
from datetime import datetime

sys.path.insert(0, "/Users/prakashtupe/lead-automation")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("pipeline")

from agents.leads_agent import LeadsAgent
from agents.scraper_agent import ScraperAgent
from agents.qualify_agent import QualifyAgent
from agents.image_agent import ImageAgent


async def draft_email_cli(brand_name, url, niche, tiktok_angle, product_count, site_description, screenshot_url):
    """Fire Claude CLI to draft personalized email with screenshot URL embedded."""
    prompt_path = os.path.join(os.path.dirname(__file__), "prompts", "email.txt")
    with open(prompt_path, "r") as f:
        template = f.read()

    prompt = template.format(
        brand_name=brand_name,
        url=url,
        niche=niche,
        tiktok_angle=tiktok_angle,
        product_count=product_count,
        site_description=site_description,
    )

    # Add instruction to embed the screenshot URL
    prompt += f"""

IMPORTANT: The screenshot of their website is hosted at this URL: {screenshot_url}
After the line "This is still your website, right?" add this HTML image tag:
<img src="{screenshot_url}" alt="{brand_name} website" style="max-width:100%;border-radius:8px;margin:10px 0;" />
"""

    proc = await asyncio.create_subprocess_exec(
        "claude", "--print", "--output-format", "json", "-p", prompt,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=120)

    if proc.returncode != 0:
        logger.error(f"Claude CLI error: {stderr.decode()[:200]}")
        return None

    response = stdout.decode().strip()
    try:
        outer = json.loads(response)
        content = outer.get("result", response)
    except json.JSONDecodeError:
        content = response

    try:
        return json.loads(content)
    except json.JSONDecodeError:
        start = content.find("{")
        end = content.rfind("}") + 1
        if start != -1 and end > start:
            try:
                return json.loads(content[start:end])
            except json.JSONDecodeError:
                pass
    return None


async def process_brand(lead):
    """Full pipeline for one brand."""
    brand = lead["brand_name"]
    logger.info(f"\n{'='*60}")
    logger.info(f"PROCESSING: {brand}")
    logger.info(f"{'='*60}")

    result = {
        "brand_name": brand,
        "url": lead["url"],
        "email": lead["email"],
        "contact_name": lead.get("contact_name", ""),
        "product_count": lead.get("product_count", ""),
        "estimated_sales": lead.get("estimated_sales", ""),
        "tiktok_followers": lead.get("tiktok_followers", 0),
        "country": lead.get("country", ""),
        "screenshot_path": "",
        "screenshot_url": "",
        "email_subject": "",
        "email_body": "",
        "niche": "",
        "qualified": False,
        "score": 0,
        "status": "pending",
    }

    # --- Step 1: Screenshot ---
    logger.info(f"[1/4] Screenshotting {brand}...")
    scrape = {"success": False, "screenshot_path": None, "scraped_data": {}}
    try:
        scrape = await ScraperAgent.run(
            brand_name=brand,
            url=lead["url"],
            product_url=lead.get("product_url", ""),
        )
        if scrape.get("screenshot_path"):
            result["screenshot_path"] = scrape["screenshot_path"]
        # Use owner name from About page if found
        owner = scrape.get("scraped_data", {}).get("owner_name", "")
        if owner and not result["contact_name"]:
            result["contact_name"] = owner
            logger.info(f"  Owner found: {owner}")
    except Exception as e:
        logger.error(f"  Screenshot failed: {e}")

    # --- Step 2: Upload screenshot to Cloudinary ---
    if result["screenshot_path"]:
        logger.info(f"[2/4] Uploading screenshot to Cloudinary...")
        screenshot_url = ImageAgent.upload(result["screenshot_path"], brand)
        if screenshot_url:
            result["screenshot_url"] = screenshot_url
            logger.info(f"  Hosted URL: {screenshot_url}")
        else:
            logger.warning(f"  Upload failed for {brand}")
    else:
        logger.warning(f"[2/4] No screenshot to upload for {brand}")

    # --- Step 3: Qualify ---
    logger.info(f"[3/4] Qualifying {brand}...")
    scraped_data = scrape.get("scraped_data", {})
    scraped_data["storeleads_product_count"] = lead.get("product_count", "")
    scraped_data["storeleads_sales"] = lead.get("estimated_sales", "")
    scraped_data["storeleads_tiktok_followers"] = lead.get("tiktok_followers", 0)
    scraped_data["storeleads_categories"] = lead.get("categories", [])

    tiktok_angle = ""
    try:
        qualification = await QualifyAgent.run(brand, lead["url"], scraped_data)
        result["qualified"] = qualification.get("qualified", False)
        result["score"] = qualification.get("score", 0)
        result["niche"] = qualification.get("niche", "ecommerce")
        tiktok_angle = qualification.get("tiktok_angle", "")
        logger.info(f"  Score: {result['score']}, Qualified: {result['qualified']}")
    except Exception as e:
        logger.error(f"  Qualification failed: {e}")
        result["niche"] = "clothing"
        tiktok_angle = "TikTok Shop can drive massive sales through creator-led content"

    # --- Step 4: Draft email (only if qualified and has screenshot URL) ---
    if (result["qualified"] or result["score"] >= 5) and result["screenshot_url"]:
        logger.info(f"[4/4] Drafting email for {brand}...")
        try:
            email_content = await draft_email_cli(
                brand_name=brand,
                url=lead["url"],
                niche=result["niche"],
                tiktok_angle=tiktok_angle,
                product_count=str(lead.get("product_count", "")),
                site_description=scraped_data.get("page_text", "")[:400],
                screenshot_url=result["screenshot_url"],
            )
            if email_content:
                result["email_subject"] = email_content.get("subject", "")
                result["email_body"] = email_content.get("body", "")
                result["status"] = "ready_to_send"
                logger.info(f"  Subject: {result['email_subject']}")
            else:
                result["status"] = "email_draft_failed"
        except Exception as e:
            logger.error(f"  Email draft failed: {e}")
            result["status"] = "email_draft_failed"
    elif not result["qualified"] and result["score"] < 5:
        result["status"] = "not_qualified"
    elif not result["screenshot_url"]:
        result["status"] = "no_screenshot"

    return result


def save_results_to_excel(results, output_path=None):
    """Save results to Excel ready for PlusVibe sending."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(
            os.path.dirname(__file__), "output", f"outreach_{timestamp}.xlsx"
        )

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()

    # === Sheet 1: Ready to Send ===
    ws = wb.active
    ws.title = "Ready to Send"

    headers = [
        "Brand Name", "To Email", "Contact Name",
        "Email Subject", "Email Body (HTML)",
        "Screenshot URL", "Niche", "Score", "Status",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="16a34a", end_color="16a34a", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    ready_count = 0
    for r in results:
        if r.get("status") == "ready_to_send":
            ready_count += 1
            ws.append([
                r.get("brand_name", ""),
                r.get("email", ""),
                r.get("contact_name", ""),
                r.get("email_subject", ""),
                r.get("email_body", ""),
                r.get("screenshot_url", ""),
                r.get("niche", ""),
                r.get("score", 0),
                r.get("status", ""),
            ])

    # === Sheet 2: All Results ===
    ws2 = wb.create_sheet("All Results")
    all_headers = [
        "Brand Name", "URL", "Email", "Contact Name",
        "Score", "Qualified", "Niche", "Status",
        "Products", "Est. Sales", "TikTok Followers", "Country",
        "Screenshot URL", "Email Subject",
    ]
    ws2.append(all_headers)

    header_fill2 = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    for col in range(1, len(all_headers) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.fill = header_fill2
        cell.font = header_font

    for r in results:
        ws2.append([
            r.get("brand_name", ""),
            r.get("url", ""),
            r.get("email", ""),
            r.get("contact_name", ""),
            r.get("score", 0),
            "Yes" if r.get("qualified") else "No",
            r.get("niche", ""),
            r.get("status", ""),
            r.get("product_count", ""),
            r.get("estimated_sales", ""),
            r.get("tiktok_followers", 0),
            r.get("country", ""),
            r.get("screenshot_url", ""),
            r.get("email_subject", ""),
        ])

    # Auto-width
    for ws_sheet in [ws, ws2]:
        for col in ws_sheet.columns:
            max_len = max(len(str(cell.value or "")[:50]) for cell in col)
            ws_sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(output_path)
    logger.info(f"[Excel] Saved {ready_count} ready-to-send emails to {output_path}")
    return output_path


async def main():
    logger.info("=" * 60)
    logger.info("FULL PIPELINE: Leads → Screenshot → Upload → Qualify → Email")
    logger.info("=" * 60)

    # Pull leads
    logger.info("\n[STEP 0] Pulling leads from Store Leads...")
    leads = LeadsAgent.pull_leads(
        query="clothing brand",
        platforms=["shopify"],
        max_leads=10,
        min_products=20,
        max_products=10000,
    )
    logger.info(f"Pulled {len(leads)} leads\n")

    if not leads:
        logger.error("No leads. Exiting.")
        return

    # Process brands (concurrency=2)
    semaphore = asyncio.Semaphore(2)
    all_results = []

    async def limited(lead):
        async with semaphore:
            return await process_brand(lead)

    results = await asyncio.gather(
        *[limited(lead) for lead in leads],
        return_exceptions=True,
    )

    for r in results:
        if isinstance(r, Exception):
            logger.error(f"Exception: {r}")
        else:
            all_results.append(r)

    # Save Excel
    output_path = save_results_to_excel(all_results)

    # Summary
    total = len(all_results)
    screenshots = sum(1 for r in all_results if r.get("screenshot_url"))
    qualified = sum(1 for r in all_results if r.get("qualified"))
    ready = sum(1 for r in all_results if r.get("status") == "ready_to_send")

    logger.info(f"\n{'='*60}")
    logger.info("PIPELINE COMPLETE")
    logger.info(f"{'='*60}")
    logger.info(f"Total brands:        {total}")
    logger.info(f"Screenshots hosted:  {screenshots}")
    logger.info(f"Qualified leads:     {qualified}")
    logger.info(f"Emails ready to send: {ready}")
    logger.info(f"Excel:               {output_path}")
    logger.info(f"{'='*60}")


if __name__ == "__main__":
    asyncio.run(main())
