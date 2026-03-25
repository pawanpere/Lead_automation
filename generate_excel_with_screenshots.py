#!/usr/bin/env python3
"""Generate Excel with actual screenshot images embedded in each row."""
import os
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "/Users/prakashtupe/lead-automation/output"
SCREENSHOTS_DIR = "/Users/prakashtupe/lead-automation/screenshots"

# Find latest pipeline Excel
excel_files = sorted(
    [f for f in os.listdir(OUTPUT_DIR) if f.startswith("outreach_") and f.endswith(".xlsx")],
    reverse=True,
)
source_path = os.path.join(OUTPUT_DIR, excel_files[0])
print(f"Reading: {source_path}")

# Read source data
wb_src = load_workbook(source_path, read_only=True)

# Read "Ready to Send" sheet
ws_src = wb_src["Ready to Send"]
ready_rows = list(ws_src.iter_rows(values_only=True))
ready_headers = ready_rows[0]
ready_data = ready_rows[1:]

# Read "All Results" for extra data
ws_all = wb_src["All Results"]
all_rows = list(ws_all.iter_rows(values_only=True))
all_data = all_rows[1:]

all_lookup = {}
for row in all_data:
    all_lookup[row[0]] = row  # key by brand name

wb_src.close()

# ============================================================
# Create new Excel with embedded screenshots
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = "Outreach with Screenshots"

# Headers
headers = [
    "Brand Name", "To Email", "Contact Name",
    "Email Subject", "Screenshot", "Email Body (HTML)",
    "Screenshot URL", "Niche", "Score",
]

# Style headers
header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)

ws.append(headers)
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Set column widths
col_widths = {
    "A": 20,   # Brand Name
    "B": 30,   # Email
    "C": 18,   # Contact Name
    "D": 35,   # Subject
    "E": 40,   # Screenshot (image column — wide)
    "F": 60,   # Email Body
    "G": 50,   # Screenshot URL
    "H": 25,   # Niche
    "I": 8,    # Score
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Add data rows with embedded screenshots
screenshot_col = 5  # Column E
row_num = 2
count = 0

for row in ready_data:
    brand_name = row[0] or ""
    to_email = row[1] or ""
    contact_name = row[2] or ""
    subject = row[3] or ""
    html_body = row[4] or ""
    screenshot_url = row[5] or ""
    niche = row[6] if len(row) > 6 else ""
    score = row[7] if len(row) > 7 else ""

    if not to_email or not subject:
        continue

    # Write data
    ws.cell(row=row_num, column=1, value=brand_name)
    ws.cell(row=row_num, column=2, value=to_email)
    ws.cell(row=row_num, column=3, value=contact_name)
    ws.cell(row=row_num, column=4, value=subject)
    # Column 5 = screenshot image (added below)
    ws.cell(row=row_num, column=6, value=html_body)
    ws.cell(row=row_num, column=7, value=screenshot_url)
    ws.cell(row=row_num, column=8, value=niche)
    ws.cell(row=row_num, column=9, value=score)

    # Embed screenshot image
    safe_name = brand_name.lower().replace(" ", "_").replace("/", "_")
    screenshot_path = os.path.join(SCREENSHOTS_DIR, f"{safe_name}.png")

    if os.path.exists(screenshot_path):
        try:
            img = Image(screenshot_path)
            # Resize to fit in the cell — ~300px wide, maintain aspect ratio
            img.width = 400
            img.height = int(400 * img.height / img.width) if img.width > 0 else 225

            # Set row height to fit the image
            ws.row_dimensions[row_num].height = max(img.height * 0.75, 180)

            # Place image in column E
            cell_ref = f"E{row_num}"
            ws.add_image(img, cell_ref)
            print(f"  ✓ {brand_name} — screenshot embedded")
        except Exception as e:
            ws.cell(row=row_num, column=5, value=f"[Error: {e}]")
            print(f"  ✗ {brand_name} — embed failed: {e}")
    else:
        ws.cell(row=row_num, column=5, value="[No screenshot]")
        print(f"  ✗ {brand_name} — no screenshot found")

    count += 1
    row_num += 1

# Save
output_path = os.path.join(OUTPUT_DIR, "outreach_with_screenshots.xlsx")
wb.save(output_path)

print(f"\nSaved: {output_path}")
print(f"Rows: {count} emails with embedded screenshots")
print(f"\nOpen the Excel → Column E has the actual screenshots visible in each row")
