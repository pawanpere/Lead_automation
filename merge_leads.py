#!/usr/bin/env python3
"""Merge Sayim's contact list with Store Leads — ALL columns from both files."""
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

SAYIM_CSV = "/Users/prakashtupe/Downloads/Sayim-Khan-s-List-2026-03-26.csv"
DOMAINS_CSV = "/Users/prakashtupe/Downloads/domains_export.csv"
OUTPUT = "/Users/prakashtupe/lead-automation/output/merged_leads.xlsx"

# --- Read Store Leads data, index by domain ---
print("Reading Store Leads export...")
domain_lookup = {}
store_headers = []
with open(DOMAINS_CSV, "r", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    store_headers = reader.fieldnames
    for row in reader:
        domain = row.get("domain", "").lower().strip().replace("www.", "")
        if domain:
            domain_lookup[domain] = row

print(f"  {len(domain_lookup)} domains, {len(store_headers)} columns")

# --- Read Sayim's contact list ---
print("Reading Sayim's contact list...")
contacts = []
sayim_headers = []
with open(SAYIM_CSV, "r", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    sayim_headers = reader.fieldnames
    for row in reader:
        contacts.append(row)

print(f"  {len(contacts)} contacts, {len(sayim_headers)} columns")

# --- Build merged headers ---
# Prefix store leads columns with "SL_" to avoid confusion
store_cols_renamed = []
for h in store_headers:
    if h == "domain":
        continue  # already have Company Domain from Sayim
    store_cols_renamed.append(("SL_" + h, h))

all_headers = list(sayim_headers) + [renamed for renamed, _ in store_cols_renamed]

print(f"  Merged columns: {len(all_headers)}")

# --- Create Excel ---
print("Merging...")
wb = Workbook()

# ---- Helper to write a sheet ----
def write_sheet(ws, title, rows_data, header_color):
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=9)

    ws.title = title
    ws.append(all_headers)
    for col in range(1, len(all_headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"

    for contact, store in rows_data:
        row = []
        # Sayim columns
        for h in sayim_headers:
            row.append(contact.get(h, ""))
        # Store Leads columns
        for renamed, original in store_cols_renamed:
            row.append(store.get(original, ""))
        ws.append(row)

    # Auto-width
    for col in ws.columns:
        max_len = 0
        for cell in col:
            max_len = max(max_len, min(len(str(cell.value or "")), 35))
        ws.column_dimensions[col[0].column_letter].width = max_len + 2


# --- Split into with/without product URL ---
with_product = []
without_product = []

for contact in contacts:
    company_domain = contact.get("Company Domain", "").lower().strip().replace("www.", "")
    store = domain_lookup.get(company_domain, {})

    product_url = store.get("most_recent_product_url", "").strip()

    if product_url:
        with_product.append((contact, store))
    else:
        without_product.append((contact, store))

print(f"  With product URL:    {len(with_product)}")
print(f"  Without product URL: {len(without_product)}")

# Sheet 1: with_product_url (green)
ws1 = wb.active
write_sheet(ws1, "with_product_url", with_product, "16a34a")

# Sheet 2: without_product_url (red)
ws2 = wb.create_sheet()
write_sheet(ws2, "without_product_url", without_product, "dc2626")

os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
wb.save(OUTPUT)

print(f"\nSaved: {OUTPUT}")
print(f"  'with_product_url':    {len(with_product)} rows")
print(f"  'without_product_url': {len(without_product)} rows")
print(f"  Total columns:         {len(all_headers)}")
